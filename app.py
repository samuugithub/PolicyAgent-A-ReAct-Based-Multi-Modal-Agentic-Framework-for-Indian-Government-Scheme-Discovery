"""
Government Scheme Assistant — Web App (Refactored)
===================================================
Clean modular architecture matching the notebook design:

  llm.py        → Groq client, token budget, field definitions, prompts
  retrieval.py  → FAISS + BM25 hybrid retrieval  (Cell 5 / 6)
  agents.py     → DocAgent → RetrievalAgent → ResponseAgent  (Cell 9 / 10 / 11)
  evaluation.py → Full Cell-17 evaluation matrix
  logger.py     → Structured pipeline logging
  app.py        → Flask routes only  ← THIS FILE
"""

import os, re, json, time, warnings, io, traceback, datetime
from typing import Any, List, Dict
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
     
load_dotenv()
warnings.filterwarnings('ignore')

# ── Modular imports ────────────────────────────────────────────────────────────
import llm                       # LLM / Groq layer
import agents                    # DocAgent → RetrievalAgent → ResponseAgent

from llm import (
    groq_call, _clean_answer, normalize_query,
    FIELD_QUESTIONS, FIELD_KEYWORDS, EXTRACT_SYSTEM,
    MODEL_FAST, MODEL_SMART, CHAT_SYSTEM,
)
from retrieval import (
    build_store, hybrid_retrieve, rerank,
    build_context, build_combined_context,
    split_sections_smart, chunk_text, get_section, keyword_slice,
)
from agents import (
    smart_chat, run_llm_agentic_workflow, run_excel_agent_loop,
    DocAgent, RetrievalAgent, ResponseAgent,
    scheme_matching_tool, excel_query_tool,
    bind_app_state,
)

try:
    from logger import PipelineLogger, log
    from evaluation import run_full_evaluation
    _modules_loaded = True
except ImportError:
    _modules_loaded = False
    def log(step, msg, level='INFO'):
        print(f'[{level}][{step}] {msg}')

# ── Flask app ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

BASE_DIR   = os.path.dirname(__file__)
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
OUTPUT_DIR = os.path.join(BASE_DIR, 'outputs')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

MASTER_XLSX = os.path.join(OUTPUT_DIR, 'ALL_SCHEMES_master.xlsx')
MASTER_HTML = os.path.join(OUTPUT_DIR, 'ALL_SCHEMES_report.html')

# ── Global state ───────────────────────────────────────────────────────────────
GROQ_API_KEY   = os.environ.get('GROQ_API_KEY', '').strip()
groq_client    = None      # kept here so routes can check truthiness
emb_model      = None
reranker       = None
kw_model       = None
models_loaded  = False

master_data        = []
all_pdf_paths      = []
all_pdf_texts      = {}
all_pdf_sections   = {}
all_pdf_stores     = {}
all_extracted_pdfs = set()


def _sync_agents():
    """Push current app globals into the agents module after any state change."""
    bind_app_state({
        'master_data':      master_data,
        'all_pdf_paths':    all_pdf_paths,
        'all_pdf_texts':    all_pdf_texts,
        'all_pdf_sections': all_pdf_sections,
        'all_pdf_stores':   all_pdf_stores,
        'emb_model':        emb_model,
        'reranker':         reranker,
    })


# ── Init ───────────────────────────────────────────────────────────────────────

def init_groq(api_key=None):
    global groq_client, GROQ_API_KEY
    ok, msg = llm.init_groq(api_key or GROQ_API_KEY)
    if ok:
        groq_client = llm.groq_client   # keep local ref in sync
    return ok, msg


def init_models():
    global emb_model, reranker, kw_model, models_loaded
    if models_loaded:
        return True, 'Models already loaded'
    try:
        import torch
        from sentence_transformers import SentenceTransformer, CrossEncoder
        from keybert import KeyBERT
        device    = 'cuda' if torch.cuda.is_available() else 'cpu'
        emb_model = SentenceTransformer('sentence-transformers/multi-qa-MiniLM-L6-cos-v1', device=device)
        reranker  = CrossEncoder('cross-encoder/ms-marco-MiniLM-L-6-v2', device=device)
        kw_model  = KeyBERT(model=emb_model)
        models_loaded = True
        _sync_agents()
        return True, f'Models loaded on {device}'
    except Exception as e:
        return False, str(e)


# ── Google Drive (Cell 2 — optional) ──────────────────────────────────────────
DRIVE_FOLDER_NAME   = 'SchemeAssistant'
MASTER_XLSX_NAME    = 'ALL_SCHEMES_master.xlsx'
MASTER_HTML_NAME    = 'ALL_SCHEMES_report.html'
CLIENT_SECRETS_FILE = os.path.join(BASE_DIR, 'client_secrets.json')
TOKEN_FILE          = os.path.join(BASE_DIR, 'token.json')
SCOPES = ['https://www.googleapis.com/auth/drive']

_drive_service   = None
_drive_folder_id = None
_drive_file_id   = None


def _get_drive_service():
    global _drive_service
    if _drive_service: return _drive_service
    if not os.path.exists(CLIENT_SECRETS_FILE): return None
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        creds = None
        if os.path.exists(TOKEN_FILE):
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow  = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(TOKEN_FILE, 'w') as f: f.write(creds.to_json())
        _drive_service = build('drive', 'v3', credentials=creds)
        return _drive_service
    except Exception: return None


def _get_or_create_folder():
    global _drive_folder_id
    if _drive_folder_id: return _drive_folder_id
    svc = _get_drive_service()
    if not svc: return None
    try:
        q   = f"name='{DRIVE_FOLDER_NAME}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        res = svc.files().list(q=q, fields='files(id,name)', spaces='drive').execute()
        files = res.get('files', [])
        if files:
            _drive_folder_id = files[0]['id']
        else:
            meta = {'name': DRIVE_FOLDER_NAME, 'mimeType': 'application/vnd.google-apps.folder'}
            f    = svc.files().create(body=meta, fields='id').execute()
            _drive_folder_id = f['id']
        return _drive_folder_id
    except Exception: return None


def _find_excel_in_drive(folder_id):
    global _drive_file_id
    if _drive_file_id: return _drive_file_id
    svc = _get_drive_service()
    if not svc or not folder_id: return None
    try:
        q   = f"name='{MASTER_XLSX_NAME}' and '{folder_id}' in parents and trashed=false"
        res = svc.files().list(q=q, fields='files(id,name)', spaces='drive').execute()
        files = res.get('files', [])
        if files: _drive_file_id = files[0]['id']
        return _drive_file_id
    except Exception: return None


def download_master_from_drive():
    svc = _get_drive_service()
    if not svc: return False, 'Drive not configured'
    try:
        from googleapiclient.http import MediaIoBaseDownload
        folder_id = _get_or_create_folder()
        if not folder_id: return False, 'No folder'
        file_id   = _find_excel_in_drive(folder_id)
        if not file_id: return False, 'No Excel on Drive yet'
        req = svc.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        dl  = MediaIoBaseDownload(buf, req)
        done = False
        while not done: _, done = dl.next_chunk()
        with open(MASTER_XLSX, 'wb') as f: f.write(buf.getvalue())
        return True, f'Downloaded {MASTER_XLSX_NAME}'
    except Exception as e: return False, str(e)


def _upload_file_to_drive(svc, folder_id, local_path, drive_name, mime_type):
    """Upload or update a single file on Google Drive. Returns (ok, message)."""
    from googleapiclient.http import MediaFileUpload
    if not os.path.exists(local_path):
        return False, f'{drive_name} not found locally'
    media = MediaFileUpload(local_path, mimetype=mime_type, resumable=True)
    # Check if file already exists in folder
    q   = f"name='{drive_name}' and '{folder_id}' in parents and trashed=false"
    res = svc.files().list(q=q, fields='files(id,name)', spaces='drive').execute()
    files = res.get('files', [])
    if files:
        svc.files().update(fileId=files[0]['id'], media_body=media).execute()
    else:
        meta = {'name': drive_name, 'parents': [folder_id]}
        svc.files().create(body=meta, media_body=media, fields='id').execute()
    return True, f'Synced {drive_name}'


def upload_master_to_drive():
    """Upload both the Excel and HTML report to Google Drive."""
    svc = _get_drive_service()
    if not svc: return False, 'Drive not configured'
    try:
        folder_id = _get_or_create_folder()
        if not folder_id: return False, 'Could not create/find Drive folder'

        msgs    = []
        success = True

        # Upload Excel
        if os.path.exists(MASTER_XLSX):
            ok, msg = _upload_file_to_drive(
                svc, folder_id, MASTER_XLSX, MASTER_XLSX_NAME,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            msgs.append(msg)
            if not ok: success = False
            else:
                # Keep legacy _drive_file_id in sync for download_master_from_drive
                global _drive_file_id
                _drive_file_id = None  # force re-lookup next time

        # Upload HTML report
        if os.path.exists(MASTER_HTML):
            ok, msg = _upload_file_to_drive(
                svc, folder_id, MASTER_HTML, MASTER_HTML_NAME, 'text/html'
            )
            msgs.append(msg)
            if not ok: success = False

        if not msgs:
            return False, 'No files to upload (run extraction first)'
        return success, ' | '.join(msgs)
    except Exception as e:
        return False, str(e)


def load_master_excel():
    global master_data, all_extracted_pdfs
    if not os.path.exists(MASTER_XLSX): return 0
    try:
        import pandas as pd
        df     = pd.read_excel(MASTER_XLSX, sheet_name='All Schemes')
        loaded = df.to_dict(orient='records')
        already = {str(r.get('PDF', '')).strip() for r in master_data}
        added   = 0
        for row in loaded:
            pdf_name = str(row.get('PDF', '')).strip()
            if pdf_name and pdf_name not in already:
                master_data.append(row)
                all_extracted_pdfs.add(pdf_name)
                already.add(pdf_name)
                added += 1
        _sync_agents()
        return added
    except Exception as e:
        print(f'Could not read Excel: {e}')
        return 0


# ── PDF processing (Cell 4) ───────────────────────────────────────────────────

def is_text_pdf(path, pages=3):
    try:
        import pdfplumber
        chars = 0
        with pdfplumber.open(path) as pdf:
            for pg in pdf.pages[:pages]:
                t = pg.extract_text()
                if t: chars += len(t.strip())
        return chars > 100
    except: return False


def read_text_pdf(path):
    import pdfplumber
    out = ''
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            t = pg.extract_text()
            if t and t.strip():
                t    = re.sub(r'\n\s*\d+\s*\n', '\n', t)
                out += '\n' + t.strip()
    return out


def read_scanned_pdf(path):
    from pdf2image import convert_from_path
    import pytesseract
    out  = ''
    imgs = convert_from_path(path, dpi=300)
    for i, img in enumerate(imgs):
        t = pytesseract.image_to_string(img, lang='eng', config='--psm 6')
        if t.strip(): out += f'\n[PAGE {i+1}]\n{t.strip()}\n'
    return out


def pdf_to_text(path):
    raw = read_text_pdf(path) if is_text_pdf(path) else read_scanned_pdf(path)
    raw = re.sub(r'\n\s*\d+\s*\n', '\n', raw)
    raw = re.sub(r'KARTAVYA BHAWAN.*?NEW DELHI.*?\n', '', raw, flags=re.IGNORECASE)
    raw = re.sub(r'\n{3,}', '\n\n', raw)
    raw = re.sub(r'[ \t]{2,}', ' ', raw)
    return raw.strip()


# ── Field extraction (Cell 7) ─────────────────────────────────────────────────

def extract_all_fields(pdf_path):
    fname   = os.path.basename(pdf_path)
    results = {'Scheme Name': fname.replace('.pdf', '').replace('_', ' '), 'PDF': fname}
    for field, question in FIELD_QUESTIONS.items():
        kws     = FIELD_KEYWORDS.get(field, [field.lower()])
        context = build_combined_context(
            query=question,
            pdf_path=pdf_path,
            all_pdf_texts=all_pdf_texts,
            all_pdf_sections=all_pdf_sections,
            all_pdf_stores=all_pdf_stores,
            emb_model=emb_model,
            reranker=reranker,
            field_keywords=kws[:6],
        )
        prompt  = (f'Context (from the policy document):\n{context}\n\n'
                   f'Question: {question}\nAnswer using bullet points. Only facts from the context:')
        raw     = groq_call(prompt, model=MODEL_SMART, max_tokens=350, temperature=0.05, system=EXTRACT_SYSTEM)
        results[field] = _clean_answer(raw)
        time.sleep(2.5)
    return results


# ── Excel / HTML export (Cell 8) ──────────────────────────────────────────────
FIELDS_ORDER = ['Scheme Name', 'PDF'] + list(FIELD_QUESTIONS.keys())


def merge_with_existing_excel(new_data, path=None):
    path          = path or MASTER_XLSX
    existing_rows = []
    if os.path.exists(path):
        try:
            import pandas as pd
            df            = pd.read_excel(path, sheet_name='All Schemes')
            existing_rows = df.to_dict(orient='records')
        except Exception as e:
            print(f'Could not read existing Excel ({e}). Starting fresh.')
    merged = {}
    for row in existing_rows:
        key = str(row.get('PDF', '')).strip()
        if key: merged[key] = row
    new_count = 0
    for row in new_data:
        key = str(row.get('PDF', '')).strip()
        if key and key not in merged: new_count += 1
        if key: merged[key] = row
    return list(merged.values()), new_count


def generate_master_excel(data, out_path=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    out_path = out_path or MASTER_XLSX
    if not data: return
    wb      = Workbook()
    ws      = wb.active
    ws.title = 'All Schemes'
    hdr_fill = PatternFill('solid', fgColor='1a3560')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill = PatternFill('solid', fgColor='EEF2FF')
    wht_fill = PatternFill('solid', fgColor='FFFFFF')
    thin     = Side(style='thin', color='CCCCCC')
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_al  = Alignment(wrap_text=True, vertical='top')
    ctr_al   = Alignment(horizontal='center', vertical='top', wrap_text=True)
    headers  = [f for f in FIELDS_ORDER if data and f in data[0]]
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = ctr_al; cell.border = brd
    for ri, row in enumerate(data, 2):
        fill = alt_fill if ri % 2 == 0 else wht_fill
        for ci, hdr in enumerate(headers, 1):
            val  = str(row.get(hdr, '')).replace('\u2022', '-').strip()
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.alignment = wrap_al
            cell.border = brd; cell.font = Font(size=10)
    for ci, hdr in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = {'Scheme Name': 28, 'PDF': 25}.get(hdr, 45)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Government Schemes Master Report'
    ws2['A1'].font = Font(bold=True, size=14, color='1a3560')
    ws2['A3'] = f'Generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws2['A4'] = f'Total Schemes: {len(data)}'
    ws2['A6'] = 'Scheme List:'; ws2['A6'].font = Font(bold=True)
    for i, row in enumerate(data, 7):
        ws2[f'A{i}'] = f'{i-6}. {row.get("Scheme Name", row.get("PDF", ""))}'
        ok = sum(1 for k, v in row.items()
                 if k not in ('PDF', 'Scheme Name') and str(v).strip() != 'Not available')
        ws2[f'B{i}'] = f'{ok}/{len(FIELD_QUESTIONS)} fields'
    wb.save(out_path)


def generate_master_html(data, out_path=None):
    out_path = out_path or MASTER_HTML
    if not data: return
    now       = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rows_html = ''
    for i, row in enumerate(data):
        bg    = '#f8f9ff' if i % 2 == 0 else '#ffffff'
        cells = ''
        for field in FIELDS_ORDER:
            val    = str(row.get(field, '')).replace('\n', '<br>').replace('•', '&bull;')
            cells += (f'<td style="padding:8px 12px;border:1px solid #dde;'
                      f'vertical-align:top;font-size:0.88em">{val}</td>')
        rows_html += f'<tr style="background:{bg}">{cells}</tr>\n'
    headers_html = ''.join(
        f'<th style="padding:10px 14px;background:#1a3560;color:#fff;'
        f'text-align:left;white-space:nowrap">{h}</th>' for h in FIELDS_ORDER
    )
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Government Schemes Master Report</title>
<style>
  body {{ font-family:'Segoe UI',Arial,sans-serif;margin:0;padding:20px;background:#f0f2f8;color:#222; }}
  h1   {{ color:#1a3560;border-bottom:3px solid #1a3560;padding-bottom:8px; }}
  .meta {{ color:#555;font-size:0.9em;margin-bottom:20px; }}
  .table-wrap {{ overflow-x:auto;border-radius:8px;box-shadow:0 2px 12px rgba(0,0,0,0.1); }}
  table {{ border-collapse:collapse;width:100%;background:#fff; }}
  tr:hover {{ background:#e8eeff !important; }}
  .badge {{ display:inline-block;background:#1a3560;color:#fff;border-radius:20px;
            padding:2px 10px;font-size:0.8em;margin-right:6px; }}
</style>
</head>
<body>
<h1>&#127981; Government Schemes Master Report</h1>
<div class="meta">
  Generated: <strong>{now}</strong> &nbsp;|&nbsp;
  Total schemes: <span class="badge">{len(data)}</span>
</div>
<div class="table-wrap">
<table>
<thead><tr>{headers_html}</tr></thead>
<tbody>
{rows_html}
</tbody>
</table>
</div>
</body>
</html>"""
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)


# ── Voice / Translation (Cell 14) ─────────────────────────────────────────────

def _is_romanized(text):
    letters = [c for c in text if c.isalpha()]
    if not letters: return False
    return (sum(1 for c in letters if ord(c) < 128) / len(letters)) > 0.80


def _extract_english_only(text):
    if not text: return text
    for marker in ['which translates to english as:', 'translates to english as:',
                   'in english:', 'english translation:', 'english:']:
        idx = text.lower().rfind(marker)
        if idx != -1:
            candidate = text[idx + len(marker):].strip().strip('"').split('\n')[0].strip()
            if len(candidate) > 5: return candidate
    english_lines = []
    for line in text.split('\n'):
        line = line.strip().strip('"')
        if not line: continue
        if sum(1 for c in line if ord(c) > 127) > len(line) * 0.3: continue
        if any(p in line.lower() for p in ['transliteration', 'translates to kannada',
                                            'translates to hindi', 'phonetic', 'romanized']): continue
        english_lines.append(line)
    return english_lines[-1].strip() if english_lines else text.strip()


_TRANS_SYS_TO_EN = (
    'You are a precise government document translator.\n'
    'RULES:\n'
    '1. Translate to English exactly -- preserve all legal terms, amounts, dates.\n'
    '2. Do NOT simplify, rephrase, or drop any detail.\n'
    '3. Output ONLY the translated text, nothing else.'
)


def translate_to_english(text, src_lang='kn'):
    try:
        from deep_translator import GoogleTranslator
        romanized = _is_romanized(text)
        if romanized:
            native_script = None
            try:
                lang_name   = {'kn': 'Kannada', 'hi': 'Hindi'}.get(src_lang, src_lang)
                translit_sys = ('You convert phonetically typed Indian language text into its correct native script.\n'
                                'OUTPUT RULES:\n1. Output ONLY the native script version — no English, no explanation.\n'
                                '2. Do not translate yet, just convert to the correct script.')
                raw = groq_call(f'Convert this romanized {lang_name} to {lang_name} script:\n{text}',
                                model=MODEL_FAST, max_tokens=100, temperature=0.0, system=translit_sys)
                if raw and not raw.lower().startswith('error'):
                    native_script = raw.strip().split('\n')[0].strip()
            except Exception: pass
            translate_input = native_script if native_script else text
            translate_src   = src_lang      if native_script else 'auto'
            try:
                return GoogleTranslator(source=translate_src, target='en').translate(translate_input).strip()
            except Exception: pass
            try:
                raw = groq_call(f'Translate to English accurately:\n{text}', model=MODEL_FAST, max_tokens=120,
                                temperature=0.0, system='Translate the following phonetically typed Indian language text to English. OUTPUT: one line, accurate English only.')
                if raw: return _extract_english_only(raw).strip()
            except Exception: pass
            return text
        else:
            lang_name = {'kn': 'Kannada', 'hi': 'Hindi'}.get(src_lang, src_lang)
            prompt    = f'Translate this {lang_name} text to English:\n\n{text}'
            try:
                raw = groq_call(prompt, model=MODEL_FAST, max_tokens=200, temperature=0.0, system=_TRANS_SYS_TO_EN)
                if raw and not raw.lower().startswith('error'):
                    result = _extract_english_only(raw).strip()
                    if result: return result
            except Exception: pass
            return GoogleTranslator(source=src_lang, target='en').translate(text)
    except Exception:
        return text


def translate_back(answer, tgt_lang='kn'):
    try:
        from deep_translator import GoogleTranslator
        _TRANS_SYS_FROM_EN = (
            'You are a precise government document translator.\n'
            'RULES:\n1. Translate the English text to the target language exactly.\n'
            '2. Preserve ALL amounts, dates, bullet structure, and legal terms.\n'
            '3. Do NOT add or remove any information.\n4. Output ONLY the translated text.'
        )
        lang_name = {'kn': 'Kannada', 'hi': 'Hindi'}.get(tgt_lang, tgt_lang)
        def _chunk(t):
            prompt = f'Translate this English text to {lang_name}:\n\n{t}'
            r = groq_call(prompt, model=MODEL_FAST, max_tokens=500, temperature=0.0, system=_TRANS_SYS_FROM_EN)
            return r.strip() if r and not r.lower().startswith('error') else None
        if len(answer) > 1200:
            paras  = [p for p in answer.split('\n\n') if p.strip()]
            result = '\n\n'.join(
                (_chunk(p) or GoogleTranslator(source='en', target=tgt_lang).translate(p)) for p in paras
            )
        else:
            result = _chunk(answer) or GoogleTranslator(source='en', target=tgt_lang).translate(answer)
        return result
    except Exception: return answer


def text_to_speech(text, lang='kn'):
    try:
        from gtts import gTTS
        lang_codes = {'kn': 'kn', 'hi': 'hi', 'en': 'en'}
        tts = gTTS(text=text[:500], lang=lang_codes.get(lang, 'en'), slow=False)
        buf = io.BytesIO()
        tts.write_to_fp(buf)
        buf.seek(0)
        return buf.read()
    except Exception: return None


# ── Document Recommender helpers (Cell 16) ────────────────────────────────────

_DOC_SYS = '''Extract student eligibility fields from a scanned Indian government document.
Return ONLY valid JSON. Use null for any field not present in this document.
{
  "doc_type"   : "<income_certificate|caste_certificate|marksheet|bonafide|other>",
  "name"       : "<student name or null>",
  "income"     : <annual family income INR as integer or null>,
  "category"   : "<SC|ST|OBC|General|EWS|NT|SBC or null>",
  "percentage" : <overall % or CGPA as float or null>,
  "class_grade": "<10th|12th|UG|PG or null>",
  "course"     : "<programme name or null>",
  "institution": "<school/college or null>",
  "state"      : "<state of domicile or null>",
  "district"   : "<district or null>",
  "age"        : <integer or null>,
  "year"       : "<academic year or null>"
}'''


def _extract_fields_from_doc(raw_text, filename=''):
    hint = f'Filename hint: {filename}\n\n' if filename else ''
    if len(raw_text) > 6000:
        doc_excerpt = raw_text[:5000] + '\n...\n' + raw_text[-1000:]
    else:
        doc_excerpt = raw_text
    user = hint + 'Document text:\n' + doc_excerpt
    llm._wait_if_needed(user, 400)
    try:
        resp = llm.groq_client.chat.completions.create(
            model=MODEL_SMART,
            messages=[{'role': 'system', 'content': _DOC_SYS},
                      {'role': 'user',   'content': user}],
            temperature=0.0, max_tokens=400,
        )
        llm._log_tokens(user, 400)
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r'^```json|^```|```$', '', raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        print(f'[extract_fields] Failed to parse LLM response: {e}')
        traceback.print_exc()
        return {}


def _merge_profiles(profiles):
    merged = {}
    for p in profiles:
        for k, v in p.items():
            if v not in (None, '', 'null', 'None') and merged.get(k) in (None, '', 'null', 'None', {}, None):
                merged[k] = v
    return merged


_GAP_SYS = '''You check which profile fields are missing AND required by Indian government scheme eligibility rules.
You are given:
  - student profile (some fields null)
  - doc_types already uploaded (do NOT ask for these again)
  - scheme eligibility summaries

Return ONLY valid JSON:
{
  "gaps": [
    {
      "field"    : "<field name>",
      "doc_label": "<Caste Certificate|Income Certificate|Marksheet|Bonafide>",
      "doc_type" : "<caste_certificate|income_certificate|marksheet|bonafide>",
      "reason"   : "<why this field matters, max 12 words>"
    }
  ]
}
CRITICAL RULES:
  - gaps[] must be EMPTY if all important fields are present.
  - NEVER include a doc_type that appears in already_uploaded_types.
  - Only include a gap if at least one scheme eligibility text explicitly mentions that field.
  - If category is null but NO scheme mentions SC/ST/OBC/caste -> do NOT add it.
  - Prefer empty gaps over false positives.
No markdown. No extra keys.'''


def _gap_analysis(profile, schemes, uploaded_types):
    elig_texts = '\n---\n'.join(
        f'{r.get("Scheme Name", "?")}: {str(r.get("Eligibility", ""))[:300]}'
        for r in schemes[:8]
    )
    missing = {k: v for k, v in profile.items() if v in (None, '', 'null', 'None')}
    user    = (f'Student profile:\n{json.dumps(profile, indent=2)}\n\n'
               f'Missing fields: {list(missing.keys())}\n'
               f'already_uploaded_types: {uploaded_types}\n\n'
               f'Scheme eligibility samples:\n{elig_texts}')
    llm._wait_if_needed(user, 300)
    try:
        resp = llm.groq_client.chat.completions.create(
            model=MODEL_FAST,
            messages=[{'role': 'system', 'content': _GAP_SYS},
                      {'role': 'user',   'content': user}],
            temperature=0.0, max_tokens=300,
        )
        llm._log_tokens(user, 300)
        raw    = resp.choices[0].message.content.strip()
        raw    = re.sub(r'^```json|^```|```$', '', raw, flags=re.MULTILINE).strip()
        result = json.loads(raw)
        gaps   = result.get('gaps', [])
        gaps   = [g for g in gaps if g.get('doc_type', '') not in uploaded_types]
        return gaps
    except Exception: return []


_SCORE_SYS = ('You are an Indian government scheme eligibility checker. '
              'Given a student profile and scheme details, respond ONLY as JSON: '
              '{"verdict":"MATCH","score":9,"reason":"<max 20 words>"}. '
              'verdict: MATCH | PARTIAL | NO. score: 0-10.')


def _score_scheme(row, profile):
    name    = row.get('Scheme Name', row.get('PDF', 'Unknown'))
    elig    = row.get('Eligibility', 'Not available')
    benefit = row.get('Benefits', 'Not available')
    pstr    = ', '.join(f'{k}: {v}' for k, v in profile.items()
                        if v not in (None, '', 'null', 'None') and k != 'doc_types_seen')
    user    = f'Profile: {pstr}\n\nScheme: {name}\nEligibility: {elig[:700]}\nBenefits: {benefit[:200]}'
    llm._wait_if_needed(user, 80)
    try:
        resp = llm.groq_client.chat.completions.create(
            model=MODEL_FAST,
            messages=[{'role': 'system', 'content': _SCORE_SYS},
                      {'role': 'user',   'content': user}],
            temperature=0.0, max_tokens=80,
        )
        llm._log_tokens(user, 80)
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r'^```json|^```|```$', '', raw, flags=re.MULTILINE).strip()
        r   = json.loads(raw)
        if not r: r = {'verdict': 'NO', 'score': 0, 'reason': 'Parse error'}
        r.update({'scheme_name': name, 'benefits': benefit[:350], 'eligibility': elig[:500]})
        return r
    except:
        return {'verdict': 'NO', 'score': 0, 'reason': 'Parse error',
                'scheme_name': name, 'benefits': benefit[:350], 'eligibility': elig[:500]}


def _ocr_image_via_groq(path):
    """Use Groq vision model to extract text from an image file.
    This is far more robust than Tesseract for real-world scanned documents."""
    import base64
    with open(path, 'rb') as f:
        img_data = base64.b64encode(f.read()).decode('utf-8')
    ext = path.rsplit('.', 1)[-1].lower()
    mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                'png': 'image/png', 'webp': 'image/webp'}
    mime_type = mime_map.get(ext, 'image/jpeg')

    # Use llama-4-scout-17b-16e-instruct for vision — it supports image inputs
    VISION_MODEL = 'meta-llama/llama-4-scout-17b-16e-instruct'
    system_prompt = (
        'You are an OCR assistant. Extract ALL text from the provided image exactly as it appears. '
        'Include every field name, value, date, number, name, and label you can read. '
        'Output the raw extracted text only — no commentary, no formatting, just the text.'
    )
    resp = llm.groq_client.chat.completions.create(
        model=VISION_MODEL,
        messages=[{
            'role': 'user',
            'content': [
                {
                    'type': 'image_url',
                    'image_url': {'url': f'data:{mime_type};base64,{img_data}'}
                },
                {
                    'type': 'text',
                    'text': (
                        'Please extract ALL text visible in this Indian government document '
                        '(income certificate, caste certificate, marksheet, or similar). '
                        'Include every field, value, name, date, and number you can see.'
                    )
                }
            ]
        }],
        temperature=0.0,
        max_tokens=1500,
    )
    return resp.choices[0].message.content.strip()


def _ocr_file(path, ext):
    if ext == 'pdf':
        return pdf_to_text(path)
    elif ext in ('png', 'jpg', 'jpeg', 'webp'):
        # Primary: Use Groq vision model (much more accurate than Tesseract for real docs)
        if groq_client:
            try:
                text = _ocr_image_via_groq(path)
                if text and len(text.strip()) > 20:
                    print(f'[ocr] Groq vision extracted {len(text)} chars from {os.path.basename(path)}')
                    return text
            except Exception as e:
                print(f'[ocr] Groq vision failed ({e}), falling back to Tesseract')
        # Fallback: Tesseract OCR
        try:
            import pytesseract
            from PIL import Image
            text = pytesseract.image_to_string(Image.open(path)).strip()
            print(f'[ocr] Tesseract extracted {len(text)} chars from {os.path.basename(path)}')
            return text
        except Exception as e:
            print(f'[ocr] Tesseract also failed: {e}')
            raise ValueError(f'Could not extract text from image: {e}')
    else:
        raise ValueError(f'Unsupported format: {ext}')


# ═══════════════════════════════════════════════════════════════════════════════
# FLASK ROUTES — thin wrappers only, all logic lives in the modules above
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index(): return render_template('index.html')


@app.route('/api/status')
def status():
    groq_ok, groq_msg = init_groq() if not groq_client else (True, 'Ready')
    return jsonify({
        'groq':             bool(groq_client),
        'groq_msg':         groq_msg,
        'models_loaded':    models_loaded,
        'schemes_count':    len(master_data),
        'pdfs_indexed':     len(all_pdf_paths),
        'drive_configured': os.path.exists(CLIENT_SECRETS_FILE),
        'scheme_names':     [r.get('Scheme Name', r.get('PDF', '')) for r in master_data],
        'pdf_names':        [os.path.basename(p) for p in all_pdf_paths],
    })


@app.route('/api/set-api-key', methods=['POST'])
def route_set_api_key():
    global GROQ_API_KEY, groq_client
    data = request.get_json()
    key  = data.get('api_key', '').strip()
    if not key: return jsonify({'error': 'No API key provided'}), 400
    GROQ_API_KEY = key
    os.environ['GROQ_API_KEY'] = key
    ok, msg = init_groq(key)
    return jsonify({'success': ok, 'message': msg})


@app.route('/api/init-models', methods=['POST'])
def route_init_models():
    ok, msg = init_models()
    return jsonify({'success': ok, 'message': msg})


@app.route('/api/load-excel', methods=['POST'])
def route_load_excel():
    added = load_master_excel()
    return jsonify({'success': True, 'added': added, 'total': len(master_data),
                    'schemes': [r.get('Scheme Name', r.get('PDF', '')) for r in master_data]})


@app.route('/api/drive-sync', methods=['POST'])
def route_drive_sync():
    action = request.get_json(silent=True) or {}
    if action.get('direction') == 'download' or not action.get('direction'):
        # Default: download fresh master from Drive and reload into memory
        ok, msg = download_master_from_drive()
        if ok:
            load_master_excel()
            _sync_agents()
        return jsonify({
            'success': ok,
            'message': msg,
            'schemes_loaded': len(master_data),
            'scheme_names': [r.get('Scheme Name', r.get('PDF', '')) for r in master_data],
        })
    else:
        if master_data:
            merged, _ = merge_with_existing_excel(master_data)
            generate_master_excel(merged)
        ok, msg = upload_master_to_drive()
        return jsonify({'success': ok, 'message': msg})


@app.route('/api/upload-pdf', methods=['POST'])
def route_upload_pdf():
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    files   = request.files.getlist('files')
    results = []
    for file in files:
        if not file.filename.lower().endswith('.pdf'):
            results.append({'file': file.filename, 'status': 'skipped', 'reason': 'Not a PDF'})
            continue
        fname     = secure_filename(file.filename)
        save_path = os.path.join(UPLOAD_DIR, fname)
        file.save(save_path)
        if fname in all_extracted_pdfs or any(str(r.get('PDF', '')) == fname for r in master_data):
            results.append({'file': fname, 'status': 'skipped', 'reason': 'Already extracted'})
            continue
        if save_path in all_pdf_paths:
            results.append({'file': fname, 'status': 'skipped', 'reason': 'Already indexed this session'})
            continue
        try:
            if not models_loaded:
                ok, msg = init_models()
                if not ok:
                    results.append({'file': fname, 'status': 'error', 'reason': f'Models not loaded: {msg}'})
                    continue
            log('upload_pdf', f'Indexing {fname}')
            text = pdf_to_text(save_path)
            all_pdf_texts[save_path]    = text
            all_pdf_paths.append(save_path)
            all_pdf_sections[save_path] = split_sections_smart(text)
            all_pdf_stores[save_path]   = build_store(text, emb_model)   # ← uses retrieval.py
            _sync_agents()
            log('upload_pdf', f'Indexed {fname} — {len(text):,} chars, '
                              f'{len(all_pdf_stores[save_path]["chunks"])} chunks')
            results.append({'file': fname, 'status': 'indexed', 'chars': len(text)})
        except Exception as e:
            results.append({'file': fname, 'status': 'error', 'reason': str(e)})
    return jsonify({'results': results, 'total_indexed': len(all_pdf_paths)})


@app.route('/api/extract', methods=['POST'])
def route_extract():
    if not groq_client:
        ok, msg = init_groq()
        if not ok: return jsonify({'error': msg}), 400
    already    = {row['PDF'] for row in master_data}
    to_process = [p for p in all_pdf_paths if os.path.basename(p) not in already]
    if not to_process:
        return jsonify({'message': 'All PDFs already extracted', 'total': len(master_data)})
    extracted = []; errors = []
    for pdf_path in to_process:
        try:
            row = extract_all_fields(pdf_path)
            master_data.append(row)
            all_extracted_pdfs.add(row['PDF'])
            extracted.append(row['Scheme Name'])
        except Exception as e:
            errors.append({'file': os.path.basename(pdf_path), 'error': str(e)})
    if master_data:
        merged, _ = merge_with_existing_excel(master_data)
        generate_master_excel(merged)
        generate_master_html(merged)
        # Auto-sync to Drive if configured
        if os.path.exists(CLIENT_SECRETS_FILE):
            try:
                ok, msg = upload_master_to_drive()
                print(f'[drive] Auto-sync after extraction: {msg}')
            except Exception as e:
                print(f'[drive] Auto-sync failed (non-fatal): {e}')
    _sync_agents()
    return jsonify({'extracted': extracted, 'errors': errors, 'total': len(master_data)})


@app.route('/api/chat', methods=['POST'])
def route_chat():
    data     = request.get_json()
    question = data.get('question', '').strip()
    if not question: return jsonify({'error': 'No question provided'}), 400
    if not groq_client:
        ok, msg = init_groq()
        if not ok: return jsonify({'error': f'Groq not initialized: {msg}'}), 400
    # If no data loaded yet, try to pull from Drive automatically
    if not master_data and not all_pdf_paths:
        try:
            ok, msg = download_master_from_drive()
            if ok:
                load_master_excel()
                _sync_agents()
        except Exception:
            pass
    try:
        answer = smart_chat(question)    # ← agents.py DocAgent→RetrievalAgent→ResponseAgent
        return jsonify({'answer': answer, 'question': question})
    except Exception as e:
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/voice-query', methods=['POST'])
def route_voice_query():
    data = request.get_json()
    text = data.get('text', '').strip()
    lang = data.get('lang', 'kn')
    if not text: return jsonify({'error': 'No text provided'}), 400
    if not groq_client:
        ok, msg = init_groq()
        if not ok: return jsonify({'error': msg}), 400
    try:
        english       = translate_to_english(text, src_lang=lang) if lang != 'en' else text
        answer        = smart_chat(english)
        native_answer = translate_back(answer, tgt_lang=lang) if lang != 'en' else answer
        return jsonify({'original': text, 'english_query': english, 'english_answer': answer,
                        'native_answer': native_answer, 'lang': lang})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tts', methods=['POST'])
def route_tts():
    data  = request.get_json()
    text  = data.get('text', '')
    lang  = data.get('lang', 'en')
    audio = text_to_speech(text, lang=lang)
    if not audio: return jsonify({'error': 'TTS failed'}), 500
    return send_file(io.BytesIO(audio), mimetype='audio/mpeg',
                     as_attachment=False, download_name='response.mp3')


@app.route('/api/recommend', methods=['POST'])
def route_recommend():
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    if not groq_client:
        ok, msg = init_groq()
        if not ok: return jsonify({'error': msg}), 400
    if not master_data:
        return jsonify({'error': 'No schemes loaded. Upload and extract scheme PDFs first.'}), 400

    files          = request.files.getlist('files')
    all_profiles   = []
    uploaded_types = []

    for file in files:
        fname     = secure_filename(file.filename)
        save_path = os.path.join(UPLOAD_DIR, fname)
        file.save(save_path)
        ext = fname.rsplit('.', 1)[-1].lower()
        try:
            text    = _ocr_file(save_path, ext)
            profile = _extract_fields_from_doc(text, filename=fname)
            all_profiles.append(profile)
            dt = profile.get('doc_type', '')
            if dt and dt not in uploaded_types: uploaded_types.append(dt)
        except Exception as e:
            print(f'[recommend] Error processing {fname}: {e}')
            traceback.print_exc()

    if not all_profiles:
        return jsonify({'error': 'Could not extract data from documents'}), 400

    merged = _merge_profiles(all_profiles)
    merged['doc_types_seen'] = ', '.join(uploaded_types)
    gaps = _gap_analysis(merged, master_data, uploaded_types)

    results = []
    for row in master_data:
        results.append(_score_scheme(row, merged))
        time.sleep(0.5)
    results.sort(key=lambda x: -x.get('score', 0))
    matches = [r for r in results if r['verdict'] in ('MATCH', 'PARTIAL')]

    return jsonify({
        'profile':            merged,
        'matches':            matches[:10],
        'all_results':        results,
        'total_schemes':      len(master_data),
        'gaps':               gaps,
        'uploaded_doc_types': uploaded_types,
    })


@app.route('/api/schemes')
def route_schemes():
    return jsonify({'schemes': master_data, 'count': len(master_data)})


@app.route('/api/download-excel')
def route_download_excel():
    if not os.path.exists(MASTER_XLSX): return jsonify({'error': 'No Excel file yet'}), 404
    return send_file(MASTER_XLSX, as_attachment=True, download_name='ALL_SCHEMES_master.xlsx')


@app.route('/api/download-html')
def route_download_html():
    if not os.path.exists(MASTER_HTML): return jsonify({'error': 'No HTML report yet'}), 404
    return send_file(MASTER_HTML, as_attachment=True, download_name='ALL_SCHEMES_report.html')


@app.route('/api/eval', methods=['POST'])
def route_eval():
    """
    Runs the full Cell-17 evaluation matrix (5 dimensions):
      1. Retrieval Quality  — FAISS + BM25 Recall@3
      2. Generation Quality — Groq faithfulness + relevance
      3. Agentic Reasoning  — ReAct loop 6-check analysis
      4. Document Extraction— OCR → structured fields
      5. End-to-End Latency — query benchmark
    """
    if not _modules_loaded:
        return jsonify({'error': 'evaluation.py not found'}), 500
    if not groq_client:
        ok, msg = init_groq()
        if not ok: return jsonify({'error': f'Groq not initialised: {msg}'}), 400
    try:
        report = run_full_evaluation(
            all_pdf_stores=all_pdf_stores,
            emb_model=emb_model,
            groq_client=groq_client,
            model_fast=MODEL_FAST,
            model_smart=MODEL_SMART,
            extract_fn=_extract_fields_from_doc,
            smart_chat_fn=smart_chat if (all_pdf_paths or master_data) else None,
        )
        return jsonify(report)
    except Exception as e:
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/logs')
def route_logs():
    if not _modules_loaded:
        return jsonify({'logs': [], 'note': 'logger.py module not loaded'})
    n = int(request.args.get('n', 200))
    return jsonify({'logs': PipelineLogger.get_recent(n)})


# ── Startup ────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_groq()
    # Always try to pull the latest master Excel from Google Drive first
    print('[startup] Attempting to download master Excel from Google Drive...')
    try:
        ok, msg = download_master_from_drive()
        print(f'[startup] Drive sync: {msg}')
    except Exception as e:
        ok = False
        print(f'[startup] Drive sync failed: {e}')
    # Fall back to local file if Drive download failed
    if os.path.exists(MASTER_XLSX):
        load_master_excel()
        _sync_agents()
    print('=' * 60)
    print('  Government Scheme Assistant — Web Server')
    print('  Architecture: llm → retrieval → agents → app')
    print('=' * 60)
    print(f'  Open:           http://localhost:5000')
    print(f'  Schemes loaded: {len(master_data)}')
    print(f'  Groq ready:     {bool(groq_client)}')
    print('=' * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)
